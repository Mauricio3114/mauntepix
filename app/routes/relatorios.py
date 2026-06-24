from datetime import date
from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.cliente_devedor import Cliente
from app.models.emprestimo import Emprestimo
from app.models.parcela import Parcela
from app.models.pagamento import Pagamento


relatorios_bp = Blueprint("relatorios", __name__)


@relatorios_bp.route("/relatorios")
@login_required
def relatorios():
    tipo = request.args.get("tipo", "parcelas_abertas")
    hoje = date.today()

    clientes = Cliente.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(Cliente.nome.asc()).all()

    total_clientes = len(clientes)

    total_parcelamentos = Emprestimo.query.filter_by(
        empresa_id=current_user.empresa_id
    ).count()

    parcelas_abertas = Parcela.query.filter(
        Parcela.empresa_id == current_user.empresa_id,
        Parcela.status.in_(["aberta", "aguardando_confirmacao"])
    ).count()

    parcelas_pagas = Parcela.query.filter_by(
        empresa_id=current_user.empresa_id,
        status="paga"
    ).count()

    parcelas_atrasadas = Parcela.query.filter(
        Parcela.empresa_id == current_user.empresa_id,
        Parcela.vencimento < hoje,
        Parcela.status != "paga"
    ).count()

    dados = buscar_dados_relatorio(tipo, hoje)

    return render_template(
        "relatorios.html",
        tipo=tipo,
        dados=dados,
        total_clientes=total_clientes,
        total_parcelamentos=total_parcelamentos,
        parcelas_abertas=parcelas_abertas,
        parcelas_pagas=parcelas_pagas,
        parcelas_atrasadas=parcelas_atrasadas
    )


@relatorios_bp.route("/relatorios/exportar-excel")
@login_required
def exportar_excel():
    tipo = request.args.get("tipo", "parcelas_abertas")
    hoje = date.today()
    dados = buscar_dados_relatorio(tipo, hoje)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    titulo = nome_relatorio(tipo)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"MauntePix - {titulo}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="111827")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = montar_headers(tipo)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    row = 4

    for item in dados:
        valores = montar_linha_excel(tipo, item)

        for col, valor in enumerate(valores, start=1):
            ws.cell(row=row, column=col, value=valor)

        row += 1

    for column_cells in ws.columns:
        tamanho = 14
        letra = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                tamanho = max(tamanho, len(str(cell.value)) + 2)

        ws.column_dimensions[letra].width = tamanho

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"{tipo}_mauntepix.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def buscar_dados_relatorio(tipo, hoje):
    empresa_id = current_user.empresa_id

    if tipo == "clientes":
        return Cliente.query.filter_by(
            empresa_id=empresa_id
        ).order_by(Cliente.nome.asc()).all()

    if tipo == "parcelamentos":
        return Emprestimo.query.options(
            joinedload(Emprestimo.cliente)
        ).filter_by(
            empresa_id=empresa_id
        ).order_by(Emprestimo.data_inicio.desc()).all()

    if tipo == "parcelas_pagas":
        return Parcela.query.options(
            joinedload(Parcela.cliente),
            joinedload(Parcela.emprestimo)
        ).filter_by(
            empresa_id=empresa_id,
            status="paga"
        ).order_by(Parcela.data_pagamento.desc()).all()

    if tipo == "inadimplentes":
        return Parcela.query.options(
            joinedload(Parcela.cliente),
            joinedload(Parcela.emprestimo)
        ).filter(
            Parcela.empresa_id == empresa_id,
            Parcela.vencimento < hoje,
            Parcela.status != "paga"
        ).order_by(Parcela.vencimento.asc()).all()

    return Parcela.query.options(
        joinedload(Parcela.cliente),
        joinedload(Parcela.emprestimo)
    ).filter(
        Parcela.empresa_id == empresa_id,
        Parcela.status.in_(["aberta", "aguardando_confirmacao"])
    ).order_by(Parcela.vencimento.asc()).all()


def nome_relatorio(tipo):
    nomes = {
        "clientes": "Clientes",
        "parcelamentos": "Parcelamentos",
        "parcelas_abertas": "Parcelas em Aberto",
        "parcelas_pagas": "Parcelas Pagas",
        "inadimplentes": "Inadimplentes / Atrasadas",
    }

    return nomes.get(tipo, "Parcelas em Aberto")


def montar_headers(tipo):
    if tipo == "clientes":
        return ["Nome", "CPF", "Telefone", "WhatsApp", "Cidade", "Estado", "Ativo"]

    if tipo == "parcelamentos":
        return [
            "Cliente",
            "Valor Liberado",
            "Juros %",
            "Valor Total",
            "Qtd Parcelas",
            "Periodicidade",
            "Data Início",
            "Status"
        ]

    return [
        "Cliente",
        "Parcela",
        "Valor",
        "Vencimento",
        "Status",
        "Data Pagamento",
        "Multa",
        "Valor Atualizado"
    ]


def montar_linha_excel(tipo, item):
    if tipo == "clientes":
        return [
            item.nome,
            item.cpf,
            item.telefone,
            item.whatsapp,
            item.cidade,
            item.estado,
            "Sim" if item.ativo else "Não"
        ]

    if tipo == "parcelamentos":
        return [
            item.cliente.nome if item.cliente else "-",
            float(item.valor_liberado or 0),
            float(item.percentual_juros or 0),
            float(item.valor_total or 0),
            item.quantidade_parcelas,
            item.periodicidade,
            item.data_inicio.strftime("%d/%m/%Y") if item.data_inicio else "-",
            item.status
        ]

    return [
        item.cliente.nome if item.cliente else "-",
        item.numero,
        float(item.valor or 0),
        item.vencimento.strftime("%d/%m/%Y") if item.vencimento else "-",
        item.status,
        item.data_pagamento.strftime("%d/%m/%Y") if item.data_pagamento else "-",
        float(item.valor_multa or 0),
        float(item.valor_atualizado or item.valor or 0)
    ]